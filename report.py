"""Generate the daily Excel activity report for a single user.

generate_report(user_id, report_date) ONLY ever queries Message rows filtered
by that user_id — this is the enforcement point for "each user only ever
sees their own mailbox" on the report-download path.

The report lists only messages the user FORWARDED, and a message appears in a
given day's report based on the day it was *forwarded* (not received). All
timestamps are stored in UTC but displayed in IST (India, UTC+5:30), so the
day window and the shown times both match what the user sees in Outlook.
"""

from datetime import datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from models import Message, SessionLocal

# (column header, minimum width)
COLUMNS = [
    ("Received From", 30),
    ("Subject", 40),
    ("Received Time", 20),
    ("To Recipients", 35),
    ("CC Recipients", 35),
    ("Has Attachments", 15),
    ("Importance", 12),
    ("Forwarded (Y/N)", 14),
    ("Forwarded To", 35),
    ("Forwarded Time", 20),
    ("Conversation ID", 25),
    ("Message ID", 25),
]

MAX_COLUMN_WIDTH = 60
DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"

# Stored timestamps are naive UTC; the tenant is India-based, so reports show
# local time. IST is a fixed UTC+5:30 offset (no daylight saving).
IST_OFFSET = timedelta(hours=5, minutes=30)


def _fmt_dt(value):
    """Format a naive-UTC datetime as an IST 'YYYY-MM-DD HH:MM:SS' string."""
    return (value + IST_OFFSET).strftime(DATETIME_FORMAT) if value else ""


def generate_report(user_id: str, report_date: str) -> BytesIO:
    """Build an .xlsx workbook of user_id's received mail for report_date
    (a 'YYYY-MM-DD' string) and return it as an in-memory BytesIO buffer.
    """
    # report_date names an IST calendar day. Stored forwarded_time is UTC, so
    # shift the IST midnight boundaries back by the offset to compare in UTC.
    day_start = datetime.strptime(report_date, "%Y-%m-%d") - IST_OFFSET
    day_end = day_start + timedelta(days=1)

    db = SessionLocal()
    try:
        rows = (
            db.query(Message)
            .filter(
                Message.user_id == user_id,
                Message.direction == "received",
                Message.forwarded.is_(True),
                Message.forwarded_time >= day_start,
                Message.forwarded_time < day_end,
            )
            .order_by(Message.forwarded_time.asc())
            .all()
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = report_date

        header_font = Font(bold=True)
        for col_idx, (title, _) in enumerate(COLUMNS, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=title)
            cell.font = header_font
        sheet.freeze_panes = "A2"

        for row_idx, msg in enumerate(rows, start=2):
            received_from = (
                f"{msg.from_name} <{msg.from_email}>" if msg.from_name else (msg.from_email or "")
            )
            sheet.cell(row=row_idx, column=1, value=received_from)
            sheet.cell(row=row_idx, column=2, value=msg.subject or "")
            sheet.cell(row=row_idx, column=3, value=_fmt_dt(msg.received_datetime))
            sheet.cell(row=row_idx, column=4, value=msg.to_recipients or "")
            sheet.cell(row=row_idx, column=5, value=msg.cc_recipients or "")
            sheet.cell(row=row_idx, column=6, value="Y" if msg.has_attachments else "N")
            sheet.cell(row=row_idx, column=7, value=msg.importance or "")
            sheet.cell(row=row_idx, column=8, value="Y" if msg.forwarded else "N")
            sheet.cell(row=row_idx, column=9, value=msg.forwarded_to or "")
            sheet.cell(row=row_idx, column=10, value=_fmt_dt(msg.forwarded_time))
            sheet.cell(row=row_idx, column=11, value=msg.conversation_id or "")
            sheet.cell(row=row_idx, column=12, value=msg.message_id or "")
    finally:
        db.close()

    # Auto-size columns to their longest cell (capped so long IDs don't blow out the sheet).
    for col_idx, (title, min_width) in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        longest = len(title)
        for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            value = row[0].value
            if value:
                longest = max(longest, len(str(value)))
        sheet.column_dimensions[col_letter].width = min(max(longest + 2, min_width), MAX_COLUMN_WIDTH)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
